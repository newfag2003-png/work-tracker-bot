import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from database import get_user, get_sessions_for_period, get_balance, get_hourly_rate, get_all_users, execute_query

# Создаём папку для временных файлов
os.makedirs("temp_excel", exist_ok=True)


def create_current_month_excel(user_id: int):
    """Создать Excel за текущий месяц для сотрудника"""
    now = datetime.now()
    return create_monthly_archive_excel(user_id, now.year, now.month)


def create_monthly_archive_excel(user_id: int, year: int, month: int):
    """Создать Excel за указанный месяц для сотрудника"""
    user = get_user(user_id)
    username = user["username"] if user else f"user_{user_id}"
    
    # Получаем сессии за месяц
    start_date = f"{year}-{month:02d}-01"
    if month == 12:
        end_date = f"{year+1}-01-01"
    else:
        end_date = f"{year}-{month+1:02d}-01"
    
    sessions = execute_query("""
        SELECT * FROM work_sessions 
        WHERE user_id = ? AND start_time >= ? AND start_time < ?
        ORDER BY start_time ASC
    """, (user_id, start_date, end_date), fetch_all=True) or []
    
    # Получаем расходы за месяц
    expenses = execute_query("""
        SELECT * FROM expenses 
        WHERE user_id = ? AND created_at >= ? AND created_at < ?
        ORDER BY created_at DESC
    """, (user_id, start_date, end_date), fetch_all=True) or []
    
    # Получаем выплаты за месяц
    payments = execute_query("""
        SELECT * FROM salary_payments 
        WHERE user_id = ? AND created_at >= ? AND created_at < ?
        ORDER BY created_at DESC
    """, (user_id, start_date, end_date), fetch_all=True) or []
    
    if not sessions:
        return None
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # ============= ЛИСТ 1: СВОДКА =============
    ws1 = wb.create_sheet("Сводка")
    ws1["A1"] = f"ОТЧЁТ ЗА {month:02d}.{year}"
    ws1["A1"].font = Font(size=16, bold=True)
    ws1.merge_cells("A1:B1")
    
    ws1["A3"] = "Сотрудник:"
    ws1["B3"] = username
    ws1["A4"] = "Период:"
    ws1["B4"] = f"{month:02d}.{year}"
    ws1["A5"] = "Дата отчёта:"
    ws1["B5"] = datetime.now().strftime("%d.%m.%Y %H:%M")
    
    total_hours = sum(s["duration"] for s in sessions) / 3600
    total_earned = sum(s["earnings"] for s in sessions)
    total_expenses = sum(e["amount"] for e in expenses if e["status"] == "approved")
    total_paid = sum(p["amount"] for p in payments if p["status"] == "confirmed")
    night_sessions = len([s for s in sessions if s["is_night"]])
    work_days = len(set(s["start_time"][:10] for s in sessions))
    
    ws1["A7"] = "ПОКАЗАТЕЛЬ"
    ws1["B7"] = "ЗНАЧЕНИЕ"
    ws1["A7"].font = Font(bold=True)
    ws1["B7"].font = Font(bold=True)
    
    rows = [
        ("Отработано дней", work_days),
        ("Всего часов", f"{total_hours:.1f} ч"),
        ("Заработано", f"{total_earned:,} ₴"),
        ("Компенсации (расходы)", f"{total_expenses:,} ₴"),
        ("Выплачено", f"{total_paid:,} ₴"),
        ("Ночных смен", night_sessions),
    ]
    
    row = 8
    for label, value in rows:
        ws1[f"A{row}"] = label
        ws1[f"B{row}"] = value
        row += 1
    
    # ============= ЛИСТ 2: РАБОЧИЕ СМЕНЫ =============
    ws2 = wb.create_sheet("Рабочие смены")
    headers = ["Дата", "Начало", "Окончание", "Объект", "Длительность", "Ночная", "Заработано", "Отчёт"]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row, session in enumerate(sessions, 2):
        ws2.cell(row=row, column=1, value=session["start_time"][:10])
        ws2.cell(row=row, column=2, value=session["start_time"][11:16])
        ws2.cell(row=row, column=3, value=session["end_time"][11:16])
        ws2.cell(row=row, column=4, value=session["object_name"])
        hours = session["duration"] / 3600
        ws2.cell(row=row, column=5, value=f"{hours:.1f} ч")
        ws2.cell(row=row, column=6, value="Да" if session["is_night"] else "Нет")
        ws2.cell(row=row, column=7, value=f"{session['earnings']:,} ₴")
        ws2.cell(row=row, column=8, value=session["daily_report"] or "—")
    
    # ============= ЛИСТ 3: РАСХОДЫ =============
    ws3 = wb.create_sheet("Расходы")
    headers = ["Дата", "Сумма", "Описание", "Статус"]
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row, expense in enumerate(expenses, 2):
        ws3.cell(row=row, column=1, value=expense["created_at"][:10] if expense["created_at"] else "")
        ws3.cell(row=row, column=2, value=f"{expense['amount']:,} ₴")
        ws3.cell(row=row, column=3, value=expense["description"])
        status = "✅ Подтверждён" if expense["status"] == "approved" else ("❌ Отклонён" if expense["status"] == "rejected" else "⏳ Ожидает")
        ws3.cell(row=row, column=4, value=status)
    
    # ============= ЛИСТ 4: ВЫПЛАТЫ =============
    ws4 = wb.create_sheet("Выплаты")
    headers = ["Дата", "Сумма", "Статус"]
    for col, header in enumerate(headers, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row, payment in enumerate(payments, 2):
        ws4.cell(row=row, column=1, value=payment["created_at"][:10] if payment["created_at"] else "")
        ws4.cell(row=row, column=2, value=f"{payment['amount']:,} ₴")
        status = "✅ Подтверждена" if payment["status"] == "confirmed" else ("❌ Отклонена" if payment["status"] == "rejected" else "⏳ Ожидает")
        ws4.cell(row=row, column=3, value=status)
    
    for col in range(1, 9):
        ws2.column_dimensions[get_column_letter(col)].width = 12
    for col in range(1, 5):
        ws3.column_dimensions[get_column_letter(col)].width = 15
        ws4.column_dimensions[get_column_letter(col)].width = 15
    
    os.makedirs(f"archives/{year}", exist_ok=True)
    filename = f"archives/{year}/{month:02d}_{username}.xlsx"
    wb.save(filename)
    return filename


def create_yearly_archive_excel(user_id: int, year: int):
    """Создать Excel за год для сотрудника (4 листа)"""
    from database import execute_query
    
    user = get_user(user_id)
    username = user["username"] if user else f"user_{user_id}"
    
    start_date = f"{year}-01-01"
    end_date = f"{year+1}-01-01"
    
    sessions = execute_query("""
        SELECT * FROM work_sessions 
        WHERE user_id = ? AND start_time >= ? AND start_time < ?
        ORDER BY start_time ASC
    """, (user_id, start_date, end_date), fetch_all=True) or []
    
    expenses = execute_query("""
        SELECT * FROM expenses 
        WHERE user_id = ? AND created_at >= ? AND created_at < ?
        ORDER BY created_at DESC
    """, (user_id, start_date, end_date), fetch_all=True) or []
    
    payments = execute_query("""
        SELECT * FROM salary_payments 
        WHERE user_id = ? AND created_at >= ? AND created_at < ?
        ORDER BY created_at DESC
    """, (user_id, start_date, end_date), fetch_all=True) or []
    
    if not sessions:
        return None
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # ============= ЛИСТ 1: СВОДКА =============
    ws1 = wb.create_sheet("Сводка")
    ws1["A1"] = f"ГОДОВОЙ ОТЧЁТ ЗА {year}"
    ws1["A1"].font = Font(size=16, bold=True)
    ws1.merge_cells("A1:B1")
    
    ws1["A3"] = "Сотрудник:"
    ws1["B3"] = username
    ws1["A4"] = "Период:"
    ws1["B4"] = f"{year} год"
    ws1["A5"] = "Дата отчёта:"
    ws1["B5"] = datetime.now().strftime("%d.%m.%Y %H:%M")
    
    total_hours = sum(s["duration"] for s in sessions) / 3600
    total_earned = sum(s["earnings"] for s in sessions)
    total_expenses = sum(e["amount"] for e in expenses if e["status"] == "approved")
    total_paid = sum(p["amount"] for p in payments if p["status"] == "confirmed")
    night_sessions = len([s for s in sessions if s["is_night"]])
    work_days = len(set(s["start_time"][:10] for s in sessions))
    
    ws1["A7"] = "ПОКАЗАТЕЛЬ"
    ws1["B7"] = "ЗНАЧЕНИЕ"
    ws1["A7"].font = Font(bold=True)
    ws1["B7"].font = Font(bold=True)
    
    rows = [
        ("Отработано дней", work_days),
        ("Всего часов", f"{total_hours:.1f} ч"),
        ("Заработано", f"{total_earned:,} ₴"),
        ("Компенсации (расходs)", f"{total_expenses:,} ₴"),
        ("Выплачено", f"{total_paid:,} ₴"),
        ("Ночных смен", night_sessions),
    ]
    
    row = 8
    for label, value in rows:
        ws1[f"A{row}"] = label
        ws1[f"B{row}"] = value
        row += 1
    
    # ============= ЛИСТ 2: РАБОЧИЕ СМЕНЫ =============
    ws2 = wb.create_sheet("Рабочие смены")
    headers = ["Дата", "Начало", "Окончание", "Объект", "Длительность", "Ночная", "Заработано", "Отчёт"]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row, session in enumerate(sessions, 2):
        ws2.cell(row=row, column=1, value=session["start_time"][:10])
        ws2.cell(row=row, column=2, value=session["start_time"][11:16])
        ws2.cell(row=row, column=3, value=session["end_time"][11:16])
        ws2.cell(row=row, column=4, value=session["object_name"])
        hours = session["duration"] / 3600
        ws2.cell(row=row, column=5, value=f"{hours:.1f} ч")
        ws2.cell(row=row, column=6, value="Да" if session["is_night"] else "Нет")
        ws2.cell(row=row, column=7, value=f"{session['earnings']:,} ₴")
        ws2.cell(row=row, column=8, value=session["daily_report"] or "—")
    
    # ============= ЛИСТ 3: РАСХОДЫ =============
    ws3 = wb.create_sheet("Расходы")
    headers = ["Дата", "Сумма", "Описание", "Статус"]
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row, expense in enumerate(expenses, 2):
        ws3.cell(row=row, column=1, value=expense["created_at"][:10] if expense["created_at"] else "")
        ws3.cell(row=row, column=2, value=f"{expense['amount']:,} ₴")
        ws3.cell(row=row, column=3, value=expense["description"])
        status = "✅ Подтверждён" if expense["status"] == "approved" else ("❌ Отклонён" if expense["status"] == "rejected" else "⏳ Ожидает")
        ws3.cell(row=row, column=4, value=status)
    
    # ============= ЛИСТ 4: ВЫПЛАТЫ =============
    ws4 = wb.create_sheet("Выплаты")
    headers = ["Дата", "Сумма", "Статус"]
    for col, header in enumerate(headers, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row, payment in enumerate(payments, 2):
        ws4.cell(row=row, column=1, value=payment["created_at"][:10] if payment["created_at"] else "")
        ws4.cell(row=row, column=2, value=f"{payment['amount']:,} ₴")
        status = "✅ Подтверждена" if payment["status"] == "confirmed" else ("❌ Отклонена" if payment["status"] == "rejected" else "⏳ Ожидает")
        ws4.cell(row=row, column=3, value=status)
    
    for col in range(1, 9):
        ws2.column_dimensions[get_column_letter(col)].width = 12
    for col in range(1, 5):
        ws3.column_dimensions[get_column_letter(col)].width = 15
        ws4.column_dimensions[get_column_letter(col)].width = 15
    
    os.makedirs(f"archives/{year}", exist_ok=True)
    filename = f"archives/{year}/{year}_годовой_{username}.xlsx"
    wb.save(filename)
    return filename


# ============= ФУНКЦИИ ДЛЯ АДМИНИСТРАТОРА (ВСЕ СОТРУДНИКИ) =============

def create_admin_current_month_excel():
    """Создать общий Excel за текущий месяц для админа (ВСЕ сотрудники)"""
    now = datetime.now()
    return create_admin_monthly_archive_excel(now.year, now.month)

def create_admin_monthly_archive_excel(year: int, month: int):
    """Создать общий Excel за указанный месяц для админа (ВСЕ сотрудники)"""
    from database import execute_query
    
    start_date = f"{year}-{month:02d}-01"
    if month == 12:
        end_date = f"{year+1}-01-01"
    else:
        end_date = f"{year}-{month+1:02d}-01"
    
    # Получаем ВСЕХ сотрудников
    users = get_all_users()
    
    if not users:
        return None
    
    # Получаем ВСЕ рабочие сессии за месяц для ВСЕХ сотрудников
    sessions = execute_query("""
        SELECT ws.*, u.username 
        FROM work_sessions ws
        INNER JOIN users u ON ws.user_id = u.user_id
        WHERE ws.start_time >= ? AND ws.start_time < ?
        ORDER BY u.username, ws.start_time ASC
    """, (start_date, end_date), fetch_all=True) or []
    
    # Получаем ВСЕ расходы за месяц для ВСЕХ сотрудников
    expenses = execute_query("""
        SELECT e.*, u.username 
        FROM expenses e
        INNER JOIN users u ON e.user_id = u.user_id
        WHERE e.created_at >= ? AND e.created_at < ?
        ORDER BY u.username, e.created_at DESC
    """, (start_date, end_date), fetch_all=True) or []
    
    # Получаем ВСЕ выплаты за месяц для ВСЕХ сотрудников
    payments = execute_query("""
        SELECT p.*, u.username 
        FROM salary_payments p
        INNER JOIN users u ON p.user_id = u.user_id
        WHERE p.created_at >= ? AND p.created_at < ?
        ORDER BY u.username, p.created_at DESC
    """, (start_date, end_date), fetch_all=True) or []
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # ============= ЛИСТ 1: СВОДКА =============
    ws1 = wb.create_sheet("Сводка")
    ws1["A1"] = f"ОБЩИЙ ОТЧЁТ ЗА {month:02d}.{year}"
    ws1["A1"].font = Font(size=16, bold=True)
    ws1.merge_cells("A1:D1")
    
    ws1["A3"] = "Дата отчёта:"
    ws1["B3"] = datetime.now().strftime("%d.%m.%Y %H:%M")
    ws1["A4"] = "Период:"
    ws1["B4"] = f"{month:02d}.{year}"
    ws1["A5"] = "Всего сотрудников:"
    ws1["B5"] = len(users)
    
    total_balance = 0
    for user in users:
        total_balance += get_balance(user["user_id"])
    
    total_hours_all = sum(s["duration"] for s in sessions) / 3600 if sessions else 0
    total_earned_all = sum(s["earnings"] for s in sessions) if sessions else 0
    total_expenses_all = sum(e["amount"] for e in expenses if e["status"] == "approved") if expenses else 0
    total_paid_all = sum(p["amount"] for p in payments if p["status"] == "confirmed") if payments else 0
    
    ws1["A7"] = "ПОКАЗАТЕЛЬ"
    ws1["B7"] = "ЗНАЧЕНИЕ"
    ws1["A7"].font = Font(bold=True)
    ws1["B7"].font = Font(bold=True)
    
    rows = [
        ("Общий баланс", f"{total_balance:,} ₴"),
        ("Всего отработано часов", f"{total_hours_all:.1f} ч"),
        ("Всего заработано", f"{total_earned_all:,} ₴"),
        ("Всего компенсаций", f"{total_expenses_all:,} ₴"),
        ("Всего выплачено", f"{total_paid_all:,} ₴"),
    ]
    
    row = 8
    for label, value in rows:
        ws1[f"A{row}"] = label
        ws1[f"B{row}"] = value
        row += 1
    
    # ============= ЛИСТ 2: СОТРУДНИКИ (СВОДКА ПО КАЖДОМУ) =============
    ws2 = wb.create_sheet("Сотрудники")
    headers = ["Имя", "Ставка", "Баланс", "Часов за месяц", "Заработано", "Компенсации", "Выплачено"]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    row = 2
    for user in users:
        user_sessions = [s for s in sessions if s["user_id"] == user["user_id"]]
        hours = sum(s["duration"] for s in user_sessions) / 3600 if user_sessions else 0
        earned = sum(s["earnings"] for s in user_sessions) if user_sessions else 0
        
        user_expenses = [e for e in expenses if e["user_id"] == user["user_id"] and e["status"] == "approved"]
        expenses_sum = sum(e["amount"] for e in user_expenses) if user_expenses else 0
        
        user_payments = [p for p in payments if p["user_id"] == user["user_id"] and p["status"] == "confirmed"]
        payments_sum = sum(p["amount"] for p in user_payments) if user_payments else 0
        
        ws2.cell(row=row, column=1, value=user["username"])
        ws2.cell(row=row, column=2, value=f"{get_hourly_rate(user['user_id'])} ₴/ч")
        ws2.cell(row=row, column=3, value=f"{get_balance(user['user_id']):,} ₴")
        ws2.cell(row=row, column=4, value=f"{hours:.1f} ч")
        ws2.cell(row=row, column=5, value=f"{earned:,} ₴")
        ws2.cell(row=row, column=6, value=f"{expenses_sum:,} ₴")
        ws2.cell(row=row, column=7, value=f"{payments_sum:,} ₴")
        row += 1
    
    for col in range(1, 8):
        ws2.column_dimensions[get_column_letter(col)].width = 16
    
    # ============= ЛИСТ 3: ВСЕ РАБОЧИЕ СМЕНЫ =============
    ws3 = wb.create_sheet("Рабочие смены (все)")
    headers = ["Сотрудник", "Дата", "Начало", "Окончание", "Объект", "Длительность", "Ночная", "Заработано", "Отчёт"]
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row, session in enumerate(sessions, 2):
        ws3.cell(row=row, column=1, value=session["username"])
        ws3.cell(row=row, column=2, value=session["start_time"][:10])
        ws3.cell(row=row, column=3, value=session["start_time"][11:16])
        ws3.cell(row=row, column=4, value=session["end_time"][11:16])
        ws3.cell(row=row, column=5, value=session["object_name"])
        hours = session["duration"] / 3600
        ws3.cell(row=row, column=6, value=f"{hours:.1f} ч")
        ws3.cell(row=row, column=7, value="Да" if session["is_night"] else "Нет")
        ws3.cell(row=row, column=8, value=f"{session['earnings']:,} ₴")
        ws3.cell(row=row, column=9, value=session["daily_report"] or "—")
    
    for col in range(1, 10):
        ws3.column_dimensions[get_column_letter(col)].width = 14
    
    # ============= ЛИСТ 4: ВСЕ РАСХОДЫ =============
    ws4 = wb.create_sheet("Расходы (все)")
    headers = ["Сотрудник", "Дата", "Сумма", "Описание", "Статус"]
    for col, header in enumerate(headers, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row, expense in enumerate(expenses, 2):
        status_text = "✅ Подтверждён" if expense["status"] == "approved" else ("❌ Отклонён" if expense["status"] == "rejected" else "⏳ Ожидает")
        ws4.cell(row=row, column=1, value=expense["username"])
        ws4.cell(row=row, column=2, value=expense["created_at"][:10] if expense["created_at"] else "")
        ws4.cell(row=row, column=3, value=f"{expense['amount']:,} ₴")
        ws4.cell(row=row, column=4, value=expense["description"])
        ws4.cell(row=row, column=5, value=status_text)
        row += 1
    
    for col in range(1, 6):
        ws4.column_dimensions[get_column_letter(col)].width = 18
    
    # ============= ЛИСТ 5: ВСЕ ВЫПЛАТЫ =============
    ws5 = wb.create_sheet("Выплаты (все)")
    headers = ["Сотрудник", "Дата", "Сумма", "Статус"]
    for col, header in enumerate(headers, 1):
        cell = ws5.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row, payment in enumerate(payments, 2):
        status_text = "✅ Подтверждена" if payment["status"] == "confirmed" else ("❌ Отклонена" if payment["status"] == "rejected" else "⏳ Ожидает")
        ws5.cell(row=row, column=1, value=payment["username"])
        ws5.cell(row=row, column=2, value=payment["created_at"][:10] if payment["created_at"] else "")
        ws5.cell(row=row, column=3, value=f"{payment['amount']:,} ₴")
        ws5.cell(row=row, column=4, value=status_text)
        row += 1
    
    for col in range(1, 5):
        ws5.column_dimensions[get_column_letter(col)].width = 18
    
    # Сохраняем файл
    os.makedirs(f"archives/{year}", exist_ok=True)
    filename = f"archives/{year}/{month:02d}_общий.xlsx"
    wb.save(filename)
    print(f"✅ Создан общий отчёт: {filename} (смен: {len(sessions)}, расходов: {len(expenses)}, выплат: {len(payments)})")
    return filename

def create_admin_yearly_archive_excel(year: int):
    """Создать общий Excel за год для админа (ВСЕ сотрудники)"""
    from database import execute_query
    
    start_date = f"{year}-01-01"
    end_date = f"{year+1}-01-01"
    
    users = get_all_users()
    
    if not users:
        return None
    
    sessions = execute_query("""
        SELECT ws.*, u.username 
        FROM work_sessions ws
        INNER JOIN users u ON ws.user_id = u.user_id
        WHERE ws.start_time >= ? AND ws.start_time < ?
        ORDER BY u.username, ws.start_time ASC
    """, (start_date, end_date), fetch_all=True) or []
    
    expenses = execute_query("""
        SELECT e.*, u.username 
        FROM expenses e
        INNER JOIN users u ON e.user_id = u.user_id
        WHERE e.created_at >= ? AND e.created_at < ?
        ORDER BY u.username, e.created_at DESC
    """, (start_date, end_date), fetch_all=True) or []
    
    payments = execute_query("""
        SELECT p.*, u.username 
        FROM salary_payments p
        INNER JOIN users u ON p.user_id = u.user_id
        WHERE p.created_at >= ? AND p.created_at < ?
        ORDER BY u.username, p.created_at DESC
    """, (start_date, end_date), fetch_all=True) or []
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Сводка
    ws1 = wb.create_sheet("Сводка")
    ws1["A1"] = f"ОБЩИЙ ГОДОВОЙ ОТЧЁТ ЗА {year}"
    ws1["A1"].font = Font(size=16, bold=True)
    ws1.merge_cells("A1:D1")
    
    ws1["A3"] = "Дата отчёта:"
    ws1["B3"] = datetime.now().strftime("%d.%m.%Y %H:%M")
    ws1["A4"] = "Период:"
    ws1["B4"] = f"{year} год"
    ws1["A5"] = "Всего сотрудников:"
    ws1["B5"] = len(users)
    
    total_balance = 0
    for user in users:
        total_balance += get_balance(user["user_id"])
    
    total_hours_all = sum(s["duration"] for s in sessions) / 3600 if sessions else 0
    total_earned_all = sum(s["earnings"] for s in sessions) if sessions else 0
    total_expenses_all = sum(e["amount"] for e in expenses if e["status"] == "approved") if expenses else 0
    total_paid_all = sum(p["amount"] for p in payments if p["status"] == "confirmed") if payments else 0
    
    ws1["A7"] = "ПОКАЗАТЕЛЬ"
    ws1["B7"] = "ЗНАЧЕНИЕ"
    ws1["A7"].font = Font(bold=True)
    ws1["B7"].font = Font(bold=True)
    
    rows = [
        ("Общий баланс", f"{total_balance:,} ₴"),
        ("Всего отработано часов", f"{total_hours_all:.1f} ч"),
        ("Всего заработано", f"{total_earned_all:,} ₴"),
        ("Всего компенсаций", f"{total_expenses_all:,} ₴"),
        ("Всего выплачено", f"{total_paid_all:,} ₴"),
    ]
    
    row = 8
    for label, value in rows:
        ws1[f"A{row}"] = label
        ws1[f"B{row}"] = value
        row += 1
    
    # Сотрудники
    ws2 = wb.create_sheet("Сотрудники")
    headers = ["Имя", "Ставка", "Баланс", "Часов за год", "Заработано", "Компенсации", "Выплачено"]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    row = 2
    for user in users:
        user_sessions = [s for s in sessions if s["user_id"] == user["user_id"]]
        hours = sum(s["duration"] for s in user_sessions) / 3600 if user_sessions else 0
        earned = sum(s["earnings"] for s in user_sessions) if user_sessions else 0
        
        user_expenses = [e for e in expenses if e["user_id"] == user["user_id"] and e["status"] == "approved"]
        expenses_sum = sum(e["amount"] for e in user_expenses) if user_expenses else 0
        
        user_payments = [p for p in payments if p["user_id"] == user["user_id"] and p["status"] == "confirmed"]
        payments_sum = sum(p["amount"] for p in user_payments) if user_payments else 0
        
        ws2.cell(row=row, column=1, value=user["username"])
        ws2.cell(row=row, column=2, value=f"{get_hourly_rate(user['user_id'])} ₴/ч")
        ws2.cell(row=row, column=3, value=f"{get_balance(user['user_id']):,} ₴")
        ws2.cell(row=row, column=4, value=f"{hours:.1f} ч")
        ws2.cell(row=row, column=5, value=f"{earned:,} ₴")
        ws2.cell(row=row, column=6, value=f"{expenses_sum:,} ₴")
        ws2.cell(row=row, column=7, value=f"{payments_sum:,} ₴")
        row += 1
    
    for col in range(1, 8):
        ws2.column_dimensions[get_column_letter(col)].width = 16
    
    # Рабочие смены (все)
    ws3 = wb.create_sheet("Рабочие смены (все)")
    headers = ["Сотрудник", "Дата", "Начало", "Окончание", "Объект", "Длительность", "Ночная", "Заработано", "Отчёт"]
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row, session in enumerate(sessions, 2):
        ws3.cell(row=row, column=1, value=session["username"])
        ws3.cell(row=row, column=2, value=session["start_time"][:10])
        ws3.cell(row=row, column=3, value=session["start_time"][11:16])
        ws3.cell(row=row, column=4, value=session["end_time"][11:16])
        ws3.cell(row=row, column=5, value=session["object_name"])
        hours = session["duration"] / 3600
        ws3.cell(row=row, column=6, value=f"{hours:.1f} ч")
        ws3.cell(row=row, column=7, value="Да" if session["is_night"] else "Нет")
        ws3.cell(row=row, column=8, value=f"{session['earnings']:,} ₴")
        ws3.cell(row=row, column=9, value=session["daily_report"] or "—")
    
    for col in range(1, 10):
        ws3.column_dimensions[get_column_letter(col)].width = 14
    
    # Расходы (все)
    ws4 = wb.create_sheet("Расходы (все)")
    headers = ["Сотрудник", "Дата", "Сумма", "Описание", "Статус"]
    for col, header in enumerate(headers, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row, expense in enumerate(expenses, 2):
        status_text = "✅ Подтверждён" if expense["status"] == "approved" else ("❌ Отклонён" if expense["status"] == "rejected" else "⏳ Ожидает")
        ws4.cell(row=row, column=1, value=expense["username"])
        ws4.cell(row=row, column=2, value=expense["created_at"][:10] if expense["created_at"] else "")
        ws4.cell(row=row, column=3, value=f"{expense['amount']:,} ₴")
        ws4.cell(row=row, column=4, value=expense["description"])
        ws4.cell(row=row, column=5, value=status_text)
        row += 1
    
    for col in range(1, 6):
        ws4.column_dimensions[get_column_letter(col)].width = 18
    
    # Выплаты (все)
    ws5 = wb.create_sheet("Выплаты (все)")
    headers = ["Сотрудник", "Дата", "Сумма", "Статус"]
    for col, header in enumerate(headers, 1):
        cell = ws5.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row, payment in enumerate(payments, 2):
        status_text = "✅ Подтверждена" if payment["status"] == "confirmed" else ("❌ Отклонена" if payment["status"] == "rejected" else "⏳ Ожидает")
        ws5.cell(row=row, column=1, value=payment["username"])
        ws5.cell(row=row, column=2, value=payment["created_at"][:10] if payment["created_at"] else "")
        ws5.cell(row=row, column=3, value=f"{payment['amount']:,} ₴")
        ws5.cell(row=row, column=4, value=status_text)
        row += 1
    
    for col in range(1, 5):
        ws5.column_dimensions[get_column_letter(col)].width = 18
    
    os.makedirs(f"archives/{year}", exist_ok=True)
    filename = f"archives/{year}/{year}_общий.xlsx"
    wb.save(filename)
    print(f"✅ Создан общий годовой отчёт: {filename}")
    return filename

def create_admin_excel(days: int = 30):
    """Создать Excel-файл для администратора за указанный период"""
    from database import execute_query
    
    users = get_all_users()
    
    if not users:
        return None
    
    cutoff = (datetime.now() - timedelta(days=days)).isoformat()
    
    # Получаем ВСЕ сессии за период
    sessions = execute_query("""
        SELECT ws.*, u.username 
        FROM work_sessions ws
        JOIN users u ON ws.user_id = u.user_id
        WHERE ws.created_at > ?
        ORDER BY u.username, ws.start_time ASC
    """, (cutoff,), fetch_all=True) or []
    
    # Получаем ВСЕ расходы за период
    expenses = execute_query("""
        SELECT e.*, u.username 
        FROM expenses e
        JOIN users u ON e.user_id = u.user_id
        WHERE e.created_at > ?
        ORDER BY u.username, e.created_at DESC
    """, (cutoff,), fetch_all=True) or []
    
    # Получаем ВСЕ выплаты за период
    payments = execute_query("""
        SELECT p.*, u.username 
        FROM salary_payments p
        JOIN users u ON p.user_id = u.user_id
        WHERE p.created_at > ?
        ORDER BY u.username, p.created_at DESC
    """, (cutoff,), fetch_all=True) or []
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # ============= ЛИСТ 1: СВОДКА =============
    ws1 = wb.create_sheet("Сводка")
    ws1["A1"] = f"ОБЩИЙ ОТЧЁТ ЗА ПОСЛЕДНИЕ {days} ДНЕЙ"
    ws1["A1"].font = Font(size=16, bold=True)
    ws1.merge_cells("A1:D1")
    
    ws1["A3"] = "Дата отчёта:"
    ws1["B3"] = datetime.now().strftime("%d.%m.%Y %H:%M")
    ws1["A4"] = "Период:"
    ws1["B4"] = f"Последние {days} дней"
    ws1["A5"] = "Всего сотрудников:"
    ws1["B5"] = len(users)
    
    total_balance = 0
    total_hours_all = sum(s["duration"] for s in sessions) / 3600 if sessions else 0
    total_earned_all = sum(s["earnings"] for s in sessions) if sessions else 0
    total_expenses_all = sum(e["amount"] for e in expenses if e["status"] == "approved") if expenses else 0
    total_paid_all = sum(p["amount"] for p in payments if p["status"] == "confirmed") if payments else 0
    
    for user in users:
        total_balance += get_balance(user["user_id"])
    
    ws1["A7"] = "ПОКАЗАТЕЛЬ"
    ws1["B7"] = "ЗНАЧЕНИЕ"
    ws1["A7"].font = Font(bold=True)
    ws1["B7"].font = Font(bold=True)
    
    rows = [
        ("Общий баланс", f"{total_balance:,} ₴"),
        ("Всего отработано часов", f"{total_hours_all:.1f} ч"),
        ("Всего заработано", f"{total_earned_all:,} ₴"),
        ("Всего компенсаций", f"{total_expenses_all:,} ₴"),
        ("Всего выплачено", f"{total_paid_all:,} ₴"),
    ]
    
    row = 8
    for label, value in rows:
        ws1[f"A{row}"] = label
        ws1[f"B{row}"] = value
        row += 1
    
    # ============= ЛИСТ 2: СОТРУДНИКИ =============
    ws2 = wb.create_sheet("Сотрудники")
    headers = ["Имя", "Ставка (₴/ч)", "Баланс (₴)", "Часов за период", "Заработано за период"]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    row = 2
    for user in users:
        user_sessions = [s for s in sessions if s["user_id"] == user["user_id"]]
        hours = sum(s["duration"] for s in user_sessions) / 3600 if user_sessions else 0
        earned = sum(s["earnings"] for s in user_sessions) if user_sessions else 0
        
        ws2.cell(row=row, column=1, value=user["username"])
        ws2.cell(row=row, column=2, value=f"{get_hourly_rate(user['user_id'])} ₴/ч")
        ws2.cell(row=row, column=3, value=f"{get_balance(user['user_id']):,} ₴")
        ws2.cell(row=row, column=4, value=f"{hours:.1f} ч")
        ws2.cell(row=row, column=5, value=f"{earned:,} ₴")
        row += 1
    
    for col in range(1, 6):
        ws2.column_dimensions[get_column_letter(col)].width = 18
    
    # ============= ЛИСТ 3: РАБОЧИЕ СМЕНЫ (ВСЕ) =============
    ws3 = wb.create_sheet("Рабочие смены")
    headers = ["Сотрудник", "Дата", "Начало", "Окончание", "Объект", "Длительность", "Ночная", "Заработано", "Отчёт"]
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row, session in enumerate(sessions, 2):
        ws3.cell(row=row, column=1, value=session["username"])
        ws3.cell(row=row, column=2, value=session["start_time"][:10])
        ws3.cell(row=row, column=3, value=session["start_time"][11:16])
        ws3.cell(row=row, column=4, value=session["end_time"][11:16])
        ws3.cell(row=row, column=5, value=session["object_name"])
        hours = session["duration"] / 3600
        ws3.cell(row=row, column=6, value=f"{hours:.1f} ч")
        ws3.cell(row=row, column=7, value="Да" if session["is_night"] else "Нет")
        ws3.cell(row=row, column=8, value=f"{session['earnings']:,} ₴")
        ws3.cell(row=row, column=9, value=session["daily_report"] or "—")
    
    for col in range(1, 10):
        ws3.column_dimensions[get_column_letter(col)].width = 14
    
    # ============= ЛИСТ 4: РАСХОДЫ (ВСЕ) =============
    ws4 = wb.create_sheet("Расходы")
    headers = ["Сотрудник", "Дата", "Сумма", "Описание", "Статус"]
    for col, header in enumerate(headers, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row, expense in enumerate(expenses, 2):
        status_text = "✅ Подтверждён" if expense["status"] == "approved" else ("❌ Отклонён" if expense["status"] == "rejected" else "⏳ Ожидает")
        ws4.cell(row=row, column=1, value=expense["username"])
        ws4.cell(row=row, column=2, value=expense["created_at"][:10] if expense["created_at"] else "")
        ws4.cell(row=row, column=3, value=f"{expense['amount']:,} ₴")
        ws4.cell(row=row, column=4, value=expense["description"])
        ws4.cell(row=row, column=5, value=status_text)
        row += 1
    
    for col in range(1, 6):
        ws4.column_dimensions[get_column_letter(col)].width = 18
    
    # ============= ЛИСТ 5: ВЫПЛАТЫ (ВСЕ) =============
    ws5 = wb.create_sheet("Выплаты")
    headers = ["Сотрудник", "Дата", "Сумма", "Статус"]
    for col, header in enumerate(headers, 1):
        cell = ws5.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row, payment in enumerate(payments, 2):
        status_text = "✅ Подтверждена" if payment["status"] == "confirmed" else ("❌ Отклонена" if payment["status"] == "rejected" else "⏳ Ожидает")
        ws5.cell(row=row, column=1, value=payment["username"])
        ws5.cell(row=row, column=2, value=payment["created_at"][:10] if payment["created_at"] else "")
        ws5.cell(row=row, column=3, value=f"{payment['amount']:,} ₴")
        ws5.cell(row=row, column=4, value=status_text)
        row += 1
    
    for col in range(1, 5):
        ws5.column_dimensions[get_column_letter(col)].width = 18
    
    filename = f"temp_excel/admin_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    os.makedirs("temp_excel", exist_ok=True)
    wb.save(filename)
    return filename